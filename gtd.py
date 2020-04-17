import re, os, time, json
from openpyxl import *

wb = Workbook()
ws = wb.create_sheet("Format_gtd", 0)
Format_gtd = wb["Format_gtd"]

Format_gtd.cell(row=1, column=1, value="Опис товара")
Format_gtd.cell(row=1, column=2, value="Кіл-ть, шт.")
Format_gtd.cell(row=1, column=3, value="Мито")
Format_gtd.cell(row=1, column=5, value="Мито, %")
Format_gtd.cell(row=1, column=6, value="Митна вартість, грн")
Format_gtd.cell(row=1, column=7, value="Сума мита")
Format_gtd.cell(row=1, column=8, value="Сума НДС")
Format_gtd.column_dimensions['A'].width = 30
Format_gtd.column_dimensions['B'].width = 15
Format_gtd.column_dimensions['D'].width = 15
Format_gtd.column_dimensions['E'].width = 20
Format_gtd.column_dimensions['F'].width = 15
Format_gtd.column_dimensions['G'].width = 15
Format_gtd.column_dimensions['H'].width = 15

# Змінна в якій зберігається назва .xlsx файла який потрібно відкрити і прочитати
newGtdFile = ""

# Скануємо папку і шукаємо .xlsx файл для відкриття і зчитування
directoryTree = os.listdir()
for fileInDirectory in directoryTree:
    filename, fileExtension = os.path.splitext(fileInDirectory)
    if (filename != "gtd") and (fileExtension == ".xlsx"):
        newGtdFile = fileInDirectory
        break

# Перевіряємо чи потрібно уточнити з якого файла брати всю необхідну інформацію
if len(newGtdFile) < 1:
    print("\nНе вдалося знайти файл!")
    newGtdFile = input("Введіть назву ГТД файла в ручну: ")
    newGtdFile = newGtdFile + ".xlsx"

# Відкриваємо файл для зчитування інформації
loadWorkbook = load_workbook(newGtdFile)
sheet = loadWorkbook.active

# Рядок з якого починаємо записувати дані у новий файл
startRow = 2
# Максимальна кількість рядків у відкритому файлі
maxRowRange = sheet.max_row - 1

# Словник з відсотками мита
toll = {}
# Значення мита у відсотках
tollValue = ''

for i in range(6, maxRowRange):
    # Отримуємо назви товарів
    nameProdact = sheet.cell(row=i, column=5).value
    # Список назв товарів з однаковою відсотковою ставкою
    listNameProdact = re.findall("[- [A-Z0-9-./]+ - [0-9]+ [шт;|м;]+]*", nameProdact)
    
    # Перебираємо список товарів і кожному вказуємо відповідний відсоток
    for x in listNameProdact:
        # Виділяємо(розділяємо) слова, назву товара і його кількість, окремо у масив
        x = x.replace(".", "")
        x = x.replace("- ", "")
        # x = x.replace(" шт", "шт")
        x = x.replace(" шт", "")
        x = x.replace(" м", "")
        x = x.strip()
        arrayOfWords = x.split(" ")

        # Записуємо назву товара у першу колонку
        Format_gtd.cell(row=startRow, column=1, value=arrayOfWords[0])
        # Записуємо кількість штук товара у другу колонку
        numberOfGoods = arrayOfWords[-1].replace(";", "")
        numberOfGoods = float(numberOfGoods)
        Format_gtd.cell(row=startRow, column=2, value=numberOfGoods)

        # Отримуємо відсотки мита та записуємо його у новий файл
        muto = sheet.cell(row=i, column=16).value
        if type(muto) == str:
            muto = muto.replace(" ", "")
            muto = muto.replace("%", "")
            muto = muto.replace(",", ".")
            muto = float(muto)
            muto = muto / 100
        muto = str(muto * 100) + " %"
        
        # Баг із отриманням відсотка
        if len(muto) > 10:
            pos = muto.find("%")
            finalString = muto[0:pos + 1]
            tollValue = finalString
            Format_gtd.cell(row=startRow, column=3, value=finalString)
        else:
            tollValue = muto
            Format_gtd.cell(row=startRow, column=3, value=muto)
        startRow = startRow + 1

    # Отримуємо значення таможньої вартості, 0 - позиція у списку
    customsValue = sheet.cell(row=i, column=15).value
    # Отримуємо значення суми мита, 1 - позиція у списку
    amountOfDuty = sheet.cell(row=i, column=17).value
    # Отримуємо значення суми НДС, 2 - позиція у списку
    amountOds = sheet.cell(row=i, column=19).value

    # Умова яка перевіряє чи існує ключ з таким значенням мита у словнику
    if tollValue in toll:
        toll[tollValue][0] = toll[tollValue][0] + customsValue
        toll[tollValue][1] = toll[tollValue][1] + amountOfDuty
        toll[tollValue][2] = toll[tollValue][2] + amountOds
    else:
        toll[tollValue] = []
        toll[tollValue].append(customsValue)
        toll[tollValue].append(amountOfDuty)
        toll[tollValue].append(amountOds)

startRowInNewDocument = 2
for b in toll:
    Format_gtd.cell(row=startRowInNewDocument, column=5, value=b)
    Format_gtd.cell(row=startRowInNewDocument, column=6, value=toll[b][0])
    Format_gtd.cell(row=startRowInNewDocument, column=7, value=toll[b][1])
    Format_gtd.cell(row=startRowInNewDocument, column=8, value=toll[b][2])
    startRowInNewDocument = startRowInNewDocument + 1

# Зберігаємо новий "gtd.xlsx" файл
wb.save("gtd.xlsx")

# Видаляємо старий не потрібний .xlsx файл
f = open("setup.json", "r")
setup = f.read()
setup = json.loads(setup)
f.close()

# Перевіряємо чи потрібно видаляти файл
if setup["deleteOldXlsxFile"]:
    os.remove(newGtdFile)
    print("\n" + newGtdFile + " - файл видалено!")

print("\nГотово!")
time.sleep(2)